"""
atualizar_programacao.py
Uso: python atualizar_programacao.py  OU  duplo clique no ATUALIZAR.bat

Estrutura esperada:
  base_preparo/*.xlsx       ← planilha de sequência de preparo (aba "CONSERVAÇÃO")
  base_fazendas/*.xlsx      ← base mestre de talhões (área por COD FAZ + TALHÃO)
  supabase_config.json      ← { "url": "...", "key": "sb_publishable_..." }

Lê a aba "CONSERVAÇÃO" (uma linha por operação, talhões agrupados em lista/faixa,
ex: "1,2,3" ou "1 AO 4"), explode cada linha em um registro por talhão, busca a
área de cada talhão na base_fazendas e faz upsert em `programacao` no Supabase —
preservando STATUS/TIPO_LINHA/CICLO/VOO já preenchidos para LAYERs existentes
(voo é mantido fora da planilha porque é atualizado pelo sync do dronemgmt e/ou
pelo ajuste manual no painel — a coluna VOOS da planilha só vale na 1ª inserção).
"""

import os
import sys

# ── Utilitários compartilhados ────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR   = os.path.dirname(_SCRIPT_DIR)   # sistema_preenchimento/
sys.path.insert(0, _SCRIPT_DIR)
from utils import layer_to_str, norm_header, mes_to_periodo, parse_talhoes, redirecionar_stdout, fechar_log

_log_fh = redirecionar_stdout(os.path.join(_BASE_DIR, 'logs', 'atualizar.log'))

import pandas as pd
import json
import glob as _glob
import requests
import openpyxl

# ── Carrega configurações ─────────────────────────────────────────────────
_cfg_path = os.path.join(_BASE_DIR, 'config.json')
try:
    with open(_cfg_path, 'r', encoding='utf-8') as _f:
        _cfg = json.load(_f)
except Exception:
    _cfg = {}

CODFAZ_EXCLUIR_PREFIXO = _cfg.get('codfaz_excluir_prefixo', '20')

_config_path = os.path.join(_BASE_DIR, 'supabase_config.json')
if not os.path.exists(_config_path):
    print(f"ERRO: Arquivo nao encontrado → {_config_path}")
    print("  Crie esse arquivo com: { \"url\": \"https://xxxx.supabase.co\", \"key\": \"sb_publishable_...\" }")
    print("  (mesma URL/key publishable já usadas em formulario.html — RLS já libera esse upsert para anon)")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

with open(_config_path, 'r', encoding='utf-8') as _f:
    _sb_cfg = json.load(_f)

SUPABASE_URL = _sb_cfg['url'].rstrip('/')
SUPABASE_KEY = _sb_cfg['key']
SB_HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
}

os.chdir(_BASE_DIR)

# ── Localiza a planilha de sequência de preparo em base_preparo/ ─────────
_preparo_found = _glob.glob("base_preparo/*.xlsx")
if not _preparo_found:
    print("ERRO: Nenhum arquivo .xlsx encontrado em base_preparo/")
    print("  Coloque a planilha de sequência de preparo na pasta base_preparo/ e tente novamente.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
SOURCE_PREPARO = _preparo_found[0]
print(f"Planilha de preparo encontrada: {SOURCE_PREPARO}")

# ── 1. Ler valores existentes no Supabase (preservar preenchimento) ──────
print("Lendo programação existente no Supabase...")
_res = requests.get(f"{SUPABASE_URL}/rest/v1/programacao?select=layer,status,tipo_linha,ciclo,voo", headers=SB_HEADERS)
if not _res.ok:
    print(f"ERRO ao ler programacao: {_res.status_code} {_res.text}")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
preserved = {}   # layer_str → (status, tipo, ciclo, voo)
for row in _res.json():
    layer = layer_to_str(row.get('layer'))
    if layer:
        preserved[layer] = (row.get('status') or '', row.get('tipo_linha') or '', row.get('ciclo') or '', row.get('voo') or '')
print(f"  {len(preserved)} linhas existentes carregadas.\n")

# ── 2. Ler base_fazendas (área por COD FAZ + TALHÃO) ──────────────────────
print("Verificando base fazendas...")
_base_faz_files = _glob.glob("base_fazendas/*.xls*")
layer_ha = {}   # (cod_faz, talhao) → area_ha
if not _base_faz_files:
    print("  AVISO: Nenhum arquivo em base_fazendas/ — área de cada talhão ficará 0.\n")
else:
    SOURCE_BASE = _base_faz_files[0]
    print(f"  Base fazendas: {SOURCE_BASE}")
    df_base = pd.read_excel(SOURCE_BASE, engine='openpyxl')
    df_base.columns = [norm_header(c) for c in df_base.columns]
    col_cod = next((c for c in ('CODIGO', 'COD FAZ', 'SECAO') if c in df_base.columns), None)
    col_tal = next((c for c in ('TALHAO', 'TALHOES') if c in df_base.columns), None)
    col_area = next((c for c in ('AREA_PROD', 'AREA_HA', 'HA', 'AREA') if c in df_base.columns), None)
    if not col_cod or not col_tal or not col_area:
        print(f"  AVISO: colunas esperadas (COD FAZ/TALHAO/AREA) não encontradas em {SOURCE_BASE} — área ficará 0.\n")
    else:
        df_base[col_cod] = pd.to_numeric(df_base[col_cod], errors='coerce')
        df_base[col_tal] = pd.to_numeric(df_base[col_tal], errors='coerce')
        df_base[col_area] = pd.to_numeric(df_base[col_area], errors='coerce')
        df_base = df_base.dropna(subset=[col_cod, col_tal])
        for _, r in df_base.iterrows():
            try:
                key = (int(r[col_cod]), int(r[col_tal]))
                layer_ha[key] = round(float(r[col_area] or 0), 2)
            except (ValueError, TypeError):
                pass
        print(f"  {len(layer_ha)} talhões na base fazendas.\n")

# ── 3. Ler aba CONSERVAÇÃO da planilha de preparo ─────────────────────────
print("Lendo planilha de preparo (aba CONSERVAÇÃO)...")
wb = openpyxl.load_workbook(SOURCE_PREPARO, data_only=True)
sheet_name = next((n for n in wb.sheetnames if norm_header(n) == 'CONSERVACAO'), None)
if not sheet_name:
    print(f"ERRO: Aba 'CONSERVAÇÃO' não encontrada. Abas disponíveis: {wb.sheetnames}")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)
ws = wb[sheet_name]

# Localiza a linha de cabeçalho (procura nas primeiras linhas por CODIGO+SECAO+TALHOES)
header_row = None
header_map = {}
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
    norms = [norm_header(c) for c in row]
    if 'CODIGO' in norms and 'TALHOES' in norms and 'SECAO' in norms:
        header_row = i
        header_map = {norms[j]: j for j in range(len(norms)) if norms[j]}
        break
if header_row is None:
    print(f"ERRO: Cabeçalho (CODIGO/SECAO/TALHOES) não encontrado na aba '{sheet_name}'.")
    fechar_log(_log_fh)
    input("\nPressione Enter para sair...")
    sys.exit(1)

idx_frente = header_map.get('FRENTE')
idx_mes    = header_map.get('MES')
idx_cod    = header_map['CODIGO']
idx_sec    = header_map['SECAO']
idx_tal    = header_map['TALHOES']
idx_tipo   = header_map.get('TIPO AREA')
idx_voo    = header_map.get('VOOS', header_map.get('VOO'))

# ── 4. Explode cada linha (lista/faixa de talhões) em um registro por talhão ──
exploded = []   # {cod_faz, fazenda, talhao, periodo_op, estagio, equipe, voo}
n_linhas = n_ignoradas = n_falha_talhao = n_admin = 0
for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), start=header_row + 1):
    cod_raw = row[idx_cod] if idx_cod < len(row) else None
    sec_raw = row[idx_sec] if idx_sec < len(row) else None
    tal_raw = row[idx_tal] if idx_tal < len(row) else None
    if cod_raw is None or sec_raw is None or tal_raw is None:
        if cod_raw is not None or sec_raw is not None or tal_raw is not None:
            n_ignoradas += 1
        continue
    try:
        cod_faz = int(cod_raw)
    except (ValueError, TypeError):
        n_ignoradas += 1
        continue
    if str(cod_faz).startswith(CODFAZ_EXCLUIR_PREFIXO):
        n_admin += 1
        continue

    talhoes = parse_talhoes(tal_raw)
    if not talhoes:
        print(f"  ⚠  Linha {i}: não foi possível interpretar TALHÕES = {tal_raw!r} (COD FAZ {cod_faz}) — ignorada.")
        n_falha_talhao += 1
        continue

    n_linhas += 1
    fazenda    = str(sec_raw).strip()
    periodo_op = mes_to_periodo(row[idx_mes]) if idx_mes is not None and idx_mes < len(row) else None
    estagio    = str(row[idx_tipo]).strip().upper() if idx_tipo is not None and idx_tipo < len(row) and row[idx_tipo] else ''
    equipe     = str(row[idx_frente]).strip().upper() if idx_frente is not None and idx_frente < len(row) and row[idx_frente] else ''
    voo        = str(row[idx_voo]).strip().upper() if idx_voo is not None and idx_voo < len(row) and row[idx_voo] else ''

    for talhao in talhoes:
        exploded.append({
            'cod_faz': cod_faz, 'fazenda': fazenda, 'talhao': talhao,
            'periodo_op': periodo_op, 'estagio': estagio, 'equipe': equipe, 'voo': voo,
        })

print(f"  {n_linhas} linha(s) de demanda, {len(exploded)} talhão(ões) explodido(s).")
if n_admin:        print(f"  Filtro administrativo (COD FAZ {CODFAZ_EXCLUIR_PREFIXO}x): {n_admin} linha(s) excluída(s).")
if n_ignoradas:     print(f"  {n_ignoradas} linha(s) parcial(is) ignorada(s) (faltava CODIGO/SECAO/TALHOES).")
if n_falha_talhao:  print(f"  {n_falha_talhao} linha(s) com TALHÕES não interpretável.")
print()

# ── 5. Monta registros por LAYER (último ganha em caso de talhão duplicado) ──
print("Montando registros para o Supabase...")
sem_area = 0
por_layer = {}
for r in exploded:
    layer_val = int(f"{r['cod_faz']}{r['talhao']:03d}")
    area_ha = layer_ha.get((r['cod_faz'], r['talhao']), 0)
    if not area_ha:
        sem_area += 1
    por_layer[layer_val] = {
        'layer': layer_val,
        'periodo_op': r['periodo_op'],
        'cod_faz': r['cod_faz'],
        'fazenda': r['fazenda'],
        'talhao': r['talhao'],
        'area_ha': area_ha,
        'estagio': r['estagio'],
        'equipe': r['equipe'],
        'voo': r['voo'],
    }
n_dup = len(exploded) - len(por_layer)
if n_dup:
    print(f"  {n_dup} talhão(ões) duplicado(s) entre linhas — mantida a última ocorrência.")
if sem_area:
    print(f"  ⚠  {sem_area} talhão(ões) sem área encontrada na base_fazendas (area_ha = 0).")

novos = 0
prog_rows = []
for layer_val, rec in por_layer.items():
    ly_str = layer_to_str(layer_val)
    if ly_str in preserved:
        status, tipo, ciclo, voo_atual = preserved[ly_str]
    else:
        status, tipo, ciclo, voo_atual = '', '', '', rec['voo']
        novos += 1
    prog_rows.append(dict(rec, status=status, tipo_linha=tipo, ciclo=ciclo, voo=voo_atual))

print(f"\nEnviando {len(prog_rows)} linhas (upsert por LAYER)...")
HEADERS_UPSERT = dict(SB_HEADERS, Prefer='resolution=merge-duplicates,return=minimal')
BATCH = 500
for i in range(0, len(prog_rows), BATCH):
    chunk = prog_rows[i:i+BATCH]
    res = requests.post(f"{SUPABASE_URL}/rest/v1/programacao", headers=HEADERS_UPSERT, json=chunk)
    if not res.ok:
        print(f"ERRO ao enviar lote {i}-{i+len(chunk)}: {res.status_code} {res.text}")
        fechar_log(_log_fh)
        input("\nPressione Enter para sair...")
        sys.exit(1)
    print(f"  {min(i+BATCH, len(prog_rows))}/{len(prog_rows)}")

print(f"\n  Upsert concluído. Novas linhas: {novos}\n")

# ── Aviso: LAYERs com preenchimento removidos desta atualização ───────────
layers_novos_str = {layer_to_str(r['layer']) for r in prog_rows}
layers_com_preenchimento = {ly for ly, vals in preserved.items() if any(str(v).strip() for v in vals)}
removidos = layers_com_preenchimento - layers_novos_str
if removidos:
    print(f"  ⚠  ATENÇÃO: {len(removidos)} LAYER(s) preenchidos não estão na nova base:")
    for ly in sorted(removidos)[:10]:
        status, tipo, ciclo, voo_atual = preserved[ly]
        print(f"     LAYER {ly}: {status} | {tipo} | {ciclo} | voo={voo_atual}")
    if len(removidos) > 10:
        print(f"     ... e mais {len(removidos)-10}")
    print("     Esses dados continuam no Supabase — apenas saíram da base atual.\n")

# ── Resumo ────────────────────────────────────────────────────────────────
print(f"{'='*50}")
print(f"  Atualizacao concluida!")
print(f"  Linhas de demanda : {n_linhas}")
print(f"  Talhões explodidos: {len(exploded)}")
print(f"  Registros enviados: {len(prog_rows)}")
print(f"  Sem área (HA=0)   : {sem_area}")
print(f"  Preservados       : {len(prog_rows) - novos}")
print(f"{'='*50}")

fechar_log(_log_fh)
input("\nPressione Enter para fechar...")
